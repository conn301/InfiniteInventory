import kivy
kivy.require('1.11.0')
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition
from kivy.uix.widget import Widget
from kivy.uix.button import Button
from kivy.properties import ObjectProperty, ListProperty
from kivy.uix.label import Label
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
import os
import pandas as pd
import xlrd
from docx import Document
import fnmatch
import openpyxl
import webbrowser
from kivy.config import Config
import sqlite3
from kivy.uix.scrollview import ScrollView
import datetime
from elasticsearch import Elasticsearch
from kivy.properties import StringProperty



#SETTING THE WINDOW OF GRAPHICAL USER INTERFACE TO A FIXED WIDTH AND HEIGHT
Config.set('graphics', 'width', '1900')
Config.set('graphics', 'height', '800')
Config.set('graphics', 'resizable',False)
database = "test3.db"


#CLASS TO HANDLE THE DATABASE FUNCTIONS
class Database(BoxLayout):

    global connect
    global c
    global data1
    data_items = ListProperty([("","","", "","","", "","", "","","","","", "","","", "")])


    connect = sqlite3.connect(database)
    c = connect.cursor()

    def createDatabase():

        #CREATING TABLE CALLED Eqlist IN THE DATABASE 'test.db'
        query_create_table = "CREATE TABLE Eqlist(date_added date, last_updated date, ip text, hostname text, maintenance text, project text, renew_mait text, ECN text, other text, GSMO text, serial_num text, model text, manufacturer text, location text, campus text, building text, desc text)"
        c.execute(query_create_table)
        connect.commit()


    def addToDatabase(df, open_wb):
        self = ""

        date_added1 = []
        last_updated1 = []
        ip1 = []
        hostname1 = []
        maintenance1 = []
        project1 = []
        renew_mait1 = []
        ECN1 = []
        other1 = []
        GSMO1 = []
        serial_num1 = []
        model1 = []
        manufacturer1 = []
        location1 = []
        campus1 = []
        building1 = []
        desc1 = []



        for i in range (1, df.nrows):

            #ADDING CELL ELEMENTS FROM EXCEL SPREADSHEET INTO DATABASE
            date_added = df.cell(i, 0).value
            if date_added != '' : #If the cell is not empty, code assumes the cell is filled with a date -- Excel saves dates as float types so it must be converted
                date_added = str(datetime.datetime(*xlrd.xldate_as_tuple(date_added, open_wb.datemode)))[0:10]
            last_updated = df.cell(i, 1).value
            if last_updated != '':
                last_updated = str(datetime.datetime(*xlrd.xldate_as_tuple(last_updated, open_wb.datemode)))[0:10] #If the cell is not empty, code assumes the cell is filled with a date -- Excel saves dates as float types so it must be converted
             #If anything inside of the cell is written outside of the date format -> (ERROR)
            ip = df.cell(i, 2).value
            hostname = df.cell(i, 3).value
            maintenance = df.cell(i,4).value
            project = df.cell(i,5).value
            renew_mait = df.cell(i,6).value
            ECN = df.cell(i,7).value
            other = df.cell(i,8).value
            GSMO = df.cell(i, 9).value
            serial_num = df.cell(i, 10).value
            model = df.cell(i, 11).value
            manufacturer = df.cell(i, 12).value
            location = df.cell(i, 13).value
            campus = df.cell(i, 14).value
            building = df.cell(i,15).value
            desc = df.cell(i,16).value
            #LAST COLUMN OF EXCEL FILE CANNOT BE EMPTY OR WILL REUSLT IN ERROR (Currently, the description column cannot be empty)
            c.execute("""INSERT INTO Eqlist (date_added, last_updated, ip, hostname, maintenance, project, renew_mait, ECN, other, GSMO, serial_num, model, manufacturer, location, campus, building, desc) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, ?)""", (date_added, last_updated, ip, hostname, maintenance, project, renew_mait, ECN, other, GSMO, serial_num, model, manufacturer, location, campus, building, desc))

            #Data from the Excel Spreadsheet is also saved in a list so it can be sent to the ElasticSearch function for searching
            text = ""
            date_added1.append(date_added)
            last_updated1.append(last_updated)
            ip1.append(ip)
            hostname1.append(hostname)
            maintenance1.append(maintenance)
            project1.append(project)
            renew_mait1.append(renew_mait)
            ECN1.append(ECN)
            other1.append(other)
            GSMO1.append(GSMO)
            serial_num1.append(serial_num)
            model1.append(model)
            manufacturer1.append(manufacturer)
            location1.append(location)
            campus1.append(campus)
            building1.append(building)
            desc1.append(desc)



        #Following the successful retrival of all data into the database -- SELECT & FROM Eqlist Displays All Items in Database
        c.execute("""SELECT * FROM Eqlist""")
        print("Inserted into Database...")
        connect.commit()
        data1 = c.fetchall()
        print(data1)
        SearchScreen.elasticsearch1(self, text, date_added1, last_updated1, ip1, hostname1, maintenance1, project1, renew_mait1, ECN1, other1, GSMO1, serial_num1, model1, manufacturer1, location1, campus1, building1, desc1)


    def viewDatabase(self):

        c.execute("SELECT * FROM Eqlist")
        self.data_items = c.fetchall()
        print(self.data_items)

    def clearDatabase(self):
        es = Elasticsearch()
        print("Testing")
        c.execute("delete from Eqlist")
        connect.commit()
        print(c.fetchall())
        es.indices.delete(index='inventory_4', ignore=[400, 404])



class LoginScreen(Screen):
    pass


class MainScreen(Screen):
    pass


class ConvertScreen(Screen):
    pass



class ViewDatabaseScreen(Screen, Database):


    def readMasterExcel(self, filename): # NOT IN USE
            global mas
            global masrows
            global mascols

            global rows, cols
            global write_to_name
            head, tail = os.path.split(str(filename))
            write_to_name = tail[:-2]
            sheet1 = pd.read_excel(write_to_name, sheet_name=0, index_col=None, header=None)
            masrows = sheet1.shape[0]
            mascols = sheet1.shape[1]
            for row in range(masrows):
                values = []
                for col in range(mascols):
                    values.append(sheet1.iloc[row, col])
                mas.append(values)
                print("g")
                print(values)

    def display(self):

        # book = xlrd.open_workbook(file_name)
        # sheet = book.sheet_by_name("Sheet1")

        # db_result = c.fetchall()
        # print(str(db_result))
        # list = []
        # for row in db_result:
        #     list.append(row)
        #
        #
        # c.close()
        #
        #
        # text = ""
        # for  i in range(len(list)):
        #     text = text + str(list[i]) + '\n'
        #self.long_text.text = (str(text))

        Database.viewDatabase(self)
      #  Database.elasticsearch1(self)


class PDFScreen(Screen, Database):



    write_to_name = "noname"

    def selected(self, filename):
        pass


    def writepdf(self, filename, path):


        #Database.createDatabase()
        #FUTURE CHANGE: Give user option to create a new database and/or check if Database already exists
        # Avoiding the error: table already exists
        if database  != "test3.db": #Change from != to == when setting up on new computer
            Database.createDatabase()


        print(filename)
        global rows, cols
        global write_to_name
        head, tail = os.path.split(str(filename))
        tail = tail[:-2]
        print(tail)
        os.chdir(path) #changes path
        open_wb = xlrd.open_workbook(tail) #file selected must be in the same folder as the project
        df = open_wb.sheet_by_name("Sheet1")
        Database.addToDatabase(df, open_wb)

        # file = pd.read_excel(filename, sheet_name=0, index_col=None, header=None)
        #rows = df.shape[0]
        # cols = df.shape[1]

class SelectFileScreen(Screen):
    write_to_name = "noname"

    def selected(self, filename):
        global rows, cols
        global write_to_name
        head, tail = os.path.split(str(filename))
        tail = tail[:-2]
        df = pd.read_excel(tail, sheet_name='Sheet1')
        # file = pd.read_excel(filename, sheet_name=0, index_col=None, header=None)
        rows = df.shape[0]
        cols = df.shape[1]
        for i in range(df.shape[0]):
            shortlist = []
            for j in df.columns:
                shortlist.append(df.loc[i, j])
            datalist.append(shortlist)

    def writeword(self):
        global datalist
        global rows
        global cols
        global write_to_name
        # write_to_name = TextInput(text='filename')
        document = Document()
        table = document.add_table(rows, cols)

        for i in range(rows):
            table_rows = table.rows[i]
            for j in range(cols):
                table_rows.cells[j].text = str(datalist[i][j])
        tempname = str(write_to_name + ".docx")
        document.save(tempname)
        datalist = []




#Unable to figure out how to display the search results on the Kivy Screen
class SearchScreen(Screen, Database):

    #
    global search_result1
    search_result1 = []
    box = ObjectProperty()

    def input(self, text):
        date_added1 = "PASSWORD"
        last_updated1 = ""
        ip1 = ""
        hostname1 = ""
        maintenance1 = ""
        project1 = ""
        renew_mait1 = ""
        ECN1 = ""
        other1 = ""
        GSMO1 = ""
        serial_num1 = ""
        model1 = ""
        manufacturer1 = ""
        location1 = ''
        campus1 = ''
        building1 = ''
        desc1 = ''

        print(str(text))
        SearchScreen.elasticsearch1(self, text, date_added1, last_updated1, ip1, hostname1, maintenance1, project1, renew_mait1, ECN1, other1, GSMO1, serial_num1, model1, manufacturer1, location1, campus1, building1, desc1)

    def result(search_result):

        text = "hi"
        date_added1 = "PASSWORD"
        last_updated1 = ""
        ip1 = ""
        hostname1 = ""
        maintenance1 = ""
        project1 = ""
        renew_mait1 = ""
        ECN1 = ""
        other1 = ""
        GSMO1 = ""
        serial_num1 = ""
        model1 = ""
        manufacturer1 = ""
        location1 = ''
        campus1 = ''
        building1 = ''
        desc1 = ''

        search_result2 = SearchScreen.elasticsearch1(self, text, date_added1, last_updated1, ip1, hostname1, maintenance1, project1, renew_mait1, ECN1, other1, GSMO1, serial_num1, model1, manufacturer1, location1, campus1, building1, desc1)
        search_result1 = search_result2
        return str(search_result1)

    def elasticsearch1(self, text, date_added1, last_updated1, ip1, hostname1, maintenance1, project1, renew_mait1, ECN1, other1, GSMO1, serial_num1, model1, manufacturer1, location1, campus1, building1, desc1):


        es = Elasticsearch()

        print("We got to this point!")

        if date_added1 != "PASSWORD":
        #Indicates that this function was called using the search box
            print("You have entered the if statement")
            import itertools
            for (item ,item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13, item14, item15, item16, item17) in zip(date_added1, last_updated1, ip1, hostname1, maintenance1, project1, renew_mait1, ECN1, other1, GSMO1, serial_num1, model1, manufacturer1, location1, campus1, building1, desc1):
                doc = {
                "Date_Added": item,
                "Last_Updated": item2,
                "IP": item3,
                "Hostname": item4,
                "Maintenance" : item5,
                "Project": item6,
                "Renew Mait" : item7,
                "ECN" : item8,
                "Other" : item9,
                "GSMO" : item10,
                "Serial Num": item11,
                "Model": item12,
                "Manufacturer" : item13,
                "Location" : item14,
                "Campus": item15,
                "Building": item16,
                "Description": item17
            }

                res = es.index(index="inventory_4", doc_type='_doc', body=doc) #maybe this doesn't send the value of res to res in the else statement
            res = es.search(index="inventory_4", scroll="1m")
            print(res)
        else:

            print("ENTERED THE SEARCH")
            res = es.search(index="inventory_4", scroll="1m")
            print(res)
            res = es.search(
                 index="inventory_4",
                 body={
                     "from": 0,
                     "size": 10,
                     "query": {
                         "multi_match": {
                             "query": text,
                             "fuzziness": "auto",
                                # "fields": [
                                #             "Date_Added",
                                #             "Last_Updated",
                                #             "IP",
                                #             "Hostname",
                                #             "Maintenance",
                                #             "Project",
                                #             "Renew Mait",
                                #             "ECN",
                                #             "Other",
                                #             "GSMO",
                                #             "Serial Num",
                                #             "Model",
                                #             "Manufacturer",
                                #             "Location",
                                #             "Campus",
                                #             "Building",
                                #             "Description",
                                #                             ]
                         }
                     }
                 }
             )
            print("Search Query" + str(res))
            print(es.indices.get_alias("*"))
            print("Got %d Result(s):" % res['hits']['total']['value'])
            search_result1 = []
            for hit in res['hits']['hits']:
                 print("Date Added: %(Date_Added)s Last Updated: %(Last_Updated)s IP: %(IP)s Hostname: %(Hostname)s Maintenance: %(Maintenance)s Project: %(Project)s Renew Mait: %(Renew Mait)s " % hit["_source"])
                 search_result = "Date Added: %(Date_Added)s Last Updated: %(Last_Updated)s IP: %(IP)s Hostname: %(Hostname)s Maintenance: %(Maintenance)s " % hit["_source"]
                 search_result1.append(search_result)


            for i in search_result1:
                self.box.add_widget(Label(text='RESULT: {0}'.format(i)))


class ScreenManagement(ScreenManager):
   pass


presentation = Builder.load_file("main.kv")


class MainApp(App):
    def build(self):
        return presentation


MainApp().run()
